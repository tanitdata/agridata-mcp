import asyncio
from tanitdata.ckan_client import CKANClient
from openpyxl import load_workbook
import io

async def process_fires(client):
    meta = await client.resource_show("ac0818b3-c726-47a2-83ab-3c5839931de6")
    data = await client.download_file(meta.get("url"))
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    fires_by_year = {}
    for row in rows[1:]:
        year = row[0]
        if isinstance(year, (int, float)) and 2015 <= int(year) <= 2025:
            try:
                num = int(row[3]) if row[3] is not None else 0
            except ValueError:
                num = 0
            fires_by_year[int(year)] = fires_by_year.get(int(year), 0) + num
    return fires_by_year

async def process_exports(client):
    meta = await client.resource_show("4c3a188e-3383-42ae-9a29-2d95cddc8955")
    data = await client.download_file(meta.get("url"))
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    exports_by_year = {}
    for row in rows[1:]:
        year = row[0]
        if isinstance(year, (int, float)) and 2015 <= int(year) <= 2025:
            total_export = sum(val for val in row[1:] if isinstance(val, (int, float)))
            exports_by_year[int(year)] = total_export
    return exports_by_year

async def main():
    client = CKANClient()
    fires = await process_fires(client)
    exports = await process_exports(client)
    await client.close()
    
    print("Year\tFires\tExports (Mille DT)")
    for year in range(2015, 2026):
        f = fires.get(year, 'N/A')
        e = exports.get(year, 'N/A')
        if isinstance(e, float):
            e = round(e, 3)
        print(f"{year}\t{f}\t{e}")

if __name__ == "__main__":
    asyncio.run(main())
