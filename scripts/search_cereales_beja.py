import asyncio
import json
import io
from tanitdata.ckan_client import CKANClient
from openpyxl import load_workbook

async def main():
    client = CKANClient()
    
    # We found df7103b7-0a85-45d6-9309-1d9967445788 has the superfice by delegetion for Beja.
    meta = await client.resource_show("df7103b7-0a85-45d6-9309-1d9967445788")
    data = await client.download_file(meta.get("url"))
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    # We want 2023. Let's find "Total Beja" or "kamel el wilaya" for 2023 or 2022/2023
    print("Looking for 2023 or 2022/2023 or 2023/2024...")
    for row in rows:
        if row[0] and ("2023" in str(row[0]) or "2022" in str(row[0])):
            if "┘â╪º┘à┘ä" in str(row[1]) or "Total" in str(row[1]) or "Beja" in str(row[1]) or "╪¿╪º╪¼╪⌐" in str(row[1]):
                print(row)
        
    await client.close()

if __name__ == "__main__":
    asyncio.run(main())
