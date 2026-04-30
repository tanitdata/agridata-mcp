"""Resource reader: download and parse non-DataStore files (CSV, XLSX)."""

from __future__ import annotations

import csv
import io
import logging
from typing import Any

from tanitdata.ckan_client import CKANClient
from tanitdata.schema_registry import SchemaRegistry
from tanitdata.utils.formatting import SKIP_COLS, format_datastore_result, format_source_footer

logger = logging.getLogger(__name__)

# Process-scoped cache: resource_id -> (fields, rows)
_cache: dict[str, tuple[list[str], list[dict[str, str]]]] = {}


def _parse_csv(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Parse CSV bytes into (field_names, rows)."""
    # Try UTF-8 first, fall back to latin-1
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    fields = list(reader.fieldnames or [])
    rows = [{k: (v or "") for k, v in row.items()} for row in reader]
    return fields, rows


def _parse_xlsx(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    """Parse XLSX bytes into (field_names, rows)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []

    rows_iter = ws.iter_rows(values_only=True)
    # First row = headers
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return [], []

    fields = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
    rows: list[dict[str, str]] = []
    for row in rows_iter:
        record = {}
        for i, val in enumerate(row):
            if i < len(fields):
                record[fields[i]] = str(val) if val is not None else ""
        rows.append(record)

    wb.close()
    return fields, rows


async def read_resource(
    client: CKANClient,
    registry: SchemaRegistry,
    resource_id: str,
    limit: int = 100,
) -> str:
    """Download and parse a non-DataStore resource file.

    Supports CSV and XLSX formats. Caches parsed results in memory.
    """
    # Check cache first
    if resource_id in _cache:
        fields, all_rows = _cache[resource_id]
        rows = all_rows[:limit]
        return _format_output(fields, rows, len(all_rows), resource_id, registry, cached=True)

    # Get resource metadata
    try:
        meta = await client.resource_show(resource_id)
    except Exception:
        meta = None
    if not meta:
        return f"Resource `{resource_id}` not found on the portal."

    # Check if DataStore-active — redirect user
    if meta.get("datastore_active"):
        return (
            f"Resource `{resource_id}` is DataStore-active. "
            f"Use `query_datastore` instead for SQL access."
        )

    url = meta.get("url", "")
    fmt = (meta.get("format") or "").upper()
    name = meta.get("name", "")

    if not url:
        return f"Resource `{resource_id}` has no download URL."

    if fmt not in ("CSV", "XLSX", "XLS"):
        return (
            f"Resource `{resource_id}` ({name}) is in **{fmt}** format. "
            f"Only CSV and XLSX files can be read. "
            f"Download manually: {url}"
        )

    # Download
    data = await client.download_file(url)
    if data is None:
        return (
            f"Resource `{resource_id}` ({name}) exceeds the 5 MB size limit. "
            f"Download manually: {url}"
        )

    # Parse
    try:
        if fmt == "CSV":
            fields, all_rows = _parse_csv(data)
        else:
            fields, all_rows = _parse_xlsx(data)
    except Exception as e:
        logger.warning("Failed to parse resource %s: %s", resource_id, e)
        return (
            f"Failed to parse resource `{resource_id}` ({name}): {e}\n"
            f"Download manually: {url}"
        )

    if not fields:
        return f"Resource `{resource_id}` ({name}) has no readable data."

    # Cache
    _cache[resource_id] = (fields, all_rows)

    rows = all_rows[:limit]
    return _format_output(fields, rows, len(all_rows), resource_id, registry, cached=False)


def _format_output(
    fields: list[str],
    rows: list[dict[str, str]],
    total: int,
    resource_id: str,
    registry: SchemaRegistry,
    cached: bool,
) -> str:
    """Format parsed file data in the same style as query_datastore."""
    clean_fields = [f for f in fields if f not in SKIP_COLS]

    lines: list[str] = []

    # Reuse the standard formatter
    lines.append(
        format_datastore_result(
            records=rows,
            fields=clean_fields,
            total=total,
            resource_id=resource_id,
        )
    )

    if cached:
        lines.append("\n*(cached — previously downloaded)*")

    # Source attribution
    source = registry.get_source_attribution(resource_id)
    if source:
        lines.append("")
        lines.append(format_source_footer([source]))

    return "\n".join(lines)
